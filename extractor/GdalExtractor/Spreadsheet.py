import csv
import io
import json
import os

import boto3
import xlrd
from openpyxl import load_workbook


class SpreadsheetExtractor:
    def __init__(self, uri, log):
        self.uri = uri
        self.spreadsheet_uri = self.uri.lstrip('s3://')
        self.log = log
        self.extractor_version = '1.0'

    def extract(self):
        self.log.info('SpreadsheetExtractor(%s)' % self.uri)

        bucket = self.spreadsheet_uri.split('/')[0]
        key = self.spreadsheet_uri.lstrip(bucket).lstrip('/')

        resp = {
            'ExtractorVersion': self.extractor_version,
            'ExtractorType': 'spreadsheet',
            'FileUri': self.uri,
            'Bucket': bucket,
            'Key': key
        }

        s3client = boto3.client(
            's3',
            region_name=os.environ.get('AWS_REG'),
            # endpoint_url='http://s3.amazonaws.com',
            config=boto3.session.Config(signature_version='s3v4')
        )

        obj = s3client.get_object(
            Bucket=bucket,
            Key=key
        )

        # All supported file extensions are not supported by one lib
        spreadsheet = obj['Body'].read()
        l_uri = self.uri.lower()
        if l_uri.endswith('.xlsx') or l_uri.endswith('.xlsm'):
            self.log.debug('extract using openpyxl - should be .xlsx or .xlsm')
            # This extraction will also trim empty rows and columns because when this lib calcs max row and col, it includes formatting
            data = self.extract_openpyxl(spreadsheet)
        elif l_uri.endswith('.xls'):
            self.log.debug('extract using xlrd - should be .xls')
            data = self.extract_xlrd(spreadsheet)
        elif l_uri.endswith('.csv'):
            self.log.debug('extract using csv - should be .csv')
            # Add file name as sheet name since csv doesn't do multiple sheets
            key_arr = key.split('/')
            data = self.extract_csv(
                spreadsheet, key_arr[len(key_arr) - 1].split('.')[0])

        resp = {**resp, **data}

        # This will convert any objects that cannot be serialized into strings (such as datetime)
        return json.loads(json.dumps(resp, default=str))

    def extract_openpyxl(self, spreadsheet):  # xlsx xlsm
        data = {}  # metadata and file contents in here
        # data_only will return the value stored the last time Excel read the sheet instead of the formula
        workbook = load_workbook(io.BytesIO(
            spreadsheet), read_only=True, data_only=True)

        # metadata
        for prop, value in vars(workbook.properties).items():
            if value is not None:
                data[prop] = value

        # file content
        sheets = workbook.get_sheet_names()
        self.log.debug('workbook sheet names: %s' % sheets)
        file_content = {}
        for sheet_name in sheets:  # one sheet at a time
            sheet = workbook.get_sheet_by_name(sheet_name)
            sheet_content = []

            # For each row, we'll look for trailing empty cells
            # The minimum number of trailing empty cells will be trimmed from all after
            trailing_empty_cells = sheet.max_column
            # Nested for to get all valid cells into a matrix (list of lists)
            for row in sheet.rows:

                # Skip completely empty rows
                if not any(cell.value for cell in row):
                    continue

                cur_row_trail_empty_cells = 0  # start at 0 to trim nothing
                row_value = []
                for cell in row:

                    # Count minimum None values that end this row
                    if cell.value is None:
                        cur_row_trail_empty_cells += 1
                    else:
                        cur_row_trail_empty_cells = 0

                    row_value.append(self.format_value(cell.value))

                sheet_content.append(row_value)
                trailing_empty_cells = min(
                    trailing_empty_cells, cur_row_trail_empty_cells)

            file_content[sheet_name] = self.trim_trailing_empty_cells(
                sheet_content, trailing_empty_cells)

        workbook.close()
        data['file_content'] = file_content
        return data

    def extract_xlrd(self, spreadsheet):  # xls
        data = {}
        workbook = xlrd.open_workbook(file_contents=spreadsheet)

        # file content, no metadata for xls files
        sheets = workbook.sheet_names()
        self.log.debug('workbook sheet names: %s' % sheets)
        file_content = {}
        for sheet_name in sheets:  # one sheet at a time
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_content = []

            # Nested for to get all valid cells into a matrix (array of arrays)
            # 0 indexed to max valid number of rows and columns
            for row_num in range(sheet.nrows):
                row = []
                for col_num in range(sheet.ncols):
                    row.append(self.format_value(
                        sheet.cell_value(row_num, col_num)))
                sheet_content.append(row)

            file_content[sheet_name] = sheet_content

        data['file_content'] = file_content
        return data

    def extract_csv(self, spreadsheet, sheet_name):  # csv
        data = {}
        workbook = csv.reader(spreadsheet.decode('utf-8'), delimiter=' ')

        # file content, csv only supports one sheet and no metadata
        file_content = {}
        file_content[sheet_name] = []
        row = []
        cell_value = ''
        for char in workbook:
            if char == [] and row != []:  # We reached the end of a row
                row.append(self.format_value(cell_value))
                cell_value = ''
                file_content[sheet_name].append(row)
                row = []
            elif char == [',']:  # We reached the end of a cell
                row.append(self.format_value(cell_value))
                cell_value = ''
            elif char != []:  # Not the end of row or cell
                cell_value = cell_value + char[0]

        data['file_content'] = file_content
        return data

    def format_value(self, string):
        # If not a string, no need to format
        if not isinstance(string, str):
            return string

        # Replace some escape charaters for readability
        new = string.replace('\t', '').replace(
            '\r', '').replace('\n', '')

        # Convert any numeric strings to numbers
        test_string = string.replace(
            '-', '', 1) if string.startswith('-') else string
        if test_string.isdigit():
            new = int(string)
        elif test_string.replace('.', '').isdigit():
            new = float(string)

        # Conform any divergent definitions of an empty cell to None
        if string == '':
            new = None

        return new

    def trim_trailing_empty_cells(self, sheet, trim_num):
        new_sheet = []
        keep_num = len(sheet[0]) - trim_num
        for row in sheet:
            new_sheet.append(row[:keep_num])
        return new_sheet

    def build_info(self, extractor_response):
        self.log.debug('spreadsheet extractor build_info')

        return {
            'ExtractorType': extractor_response['ExtractorType'],
            'ExtractorVersion': extractor_response['ExtractorVersion'],
        }
