import csv
import io
import os

import boto3
import xlrd
from openpyxl import load_workbook


class SpreadsheetExtractor:
    def __init__(self, uri, log):
        self.uri = uri
        self.spreadsheet_uri = self.uri.lstrip('s3://')
        self.log = log
        self.extractor_version = '1.0'

    def extract(self):
        self.log.info('SpreadsheetExtractor(%s)' % self.uri)

        bucket = self.spreadsheet_uri.split('/')[0]
        key = self.spreadsheet_uri.lstrip(bucket).lstrip('/')

        resp = {
            'ExtractorVersion': self.extractor_version,
            'ExtractorType': 'spreadsheet',
            'FileUri': self.uri,
            'Bucket': bucket,
            'Key': key
        }

        s3client = boto3.client(
            's3',
            region_name=os.environ.get('AWS_REG'),
            # endpoint_url='http://s3.amazonaws.com',
            config=boto3.session.Config(signature_version='s3v4')
        )

        obj = s3client.get_object(
            Bucket=bucket,
            Key=key
        )

        # All supported file extensions are not supported by one lib
        spreadsheet = obj['Body'].read()
        l_uri = self.uri.lower()
        if l_uri.endswith('.xlsx') or l_uri.endswith('.xlsm'):
            data = self.extract_openpyxl(spreadsheet)
        elif l_uri.endswith('.xls'):
            data = self.extract_xlrd(spreadsheet)
        elif l_uri.endswith('.csv'):
            data = self.extract_csv(spreadsheet)

        resp = {**resp, **data}

        return resp

    def extract_openpyxl(self, spreadsheet):
        data = {}
        workbook = load_workbook(io.BytesIO(spreadsheet), read_only=True)

        # metadata
        for prop, value in vars(workbook.properties).items():
            if value is not None:
                data[prop] = value

        # file content
        sheets = workbook.get_sheet_names()
        file_content = {}
        for sheet_name in sheets:
            sheet = workbook.get_sheet_by_name(sheet_name)
            self.log.warn(sheet_name)
            self.log.warn(sheet.max_row)
            self.log.warn(sheet.max_column)
            if (sheet.max_row * sheet.max_column) > 1000:
                continue

            sheet_content = []

            for row_num in range(1, sheet.max_row + 1):
                row = []
                for col_num in range(1, sheet.max_column + 1):
                    row.append(sheet.cell(row=row_num, column=col_num).value)
                sheet_content.append(row)
            self.log.warn(sheet_content)

            file_content[sheet_name] = sheet_content

        data['file_content'] = file_content
        return data

    def extract_xlrd(self, spreadsheet):
        data = {}
        workbook = None
        return data

    def extract_csv(self, spreadsheet):
        data = {}
        csv = None
        return data

